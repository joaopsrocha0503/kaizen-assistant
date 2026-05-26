import io
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from dotenv import load_dotenv

import database as db
import ai_service as ai

load_dotenv()

app = Flask(__name__, static_folder="static", static_url_path="/static")
CORS(app)

db.init_db()


# --- Static ---

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


# --- KPIs ---

@app.route("/api/kpis", methods=["GET"])
def kpis():
    return jsonify(db.get_kpis())


# --- Stats ---

@app.route("/api/stats/monthly", methods=["GET"])
def monthly_stats():
    months = request.args.get("months", 6, type=int)
    return jsonify(db.get_monthly_problem_counts(months=months))


# --- Problems ---

@app.route("/api/problems", methods=["GET"])
def list_problems():
    status = request.args.get("status")
    area = request.args.get("area")
    priority = request.args.get("priority")
    problems = db.list_problems(status=status, area=area, priority=priority)
    return jsonify(problems)


@app.route("/api/problems", methods=["POST"])
def create_problem():
    data = request.get_json()
    required = ["title", "description", "area", "responsible", "priority"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Campos obrigatórios em falta: {', '.join(missing)}"}), 400
    problem = db.create_problem(data)
    return jsonify(problem), 201


@app.route("/api/problems/<int:problem_id>", methods=["GET"])
def get_problem(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    return jsonify(problem)


@app.route("/api/problems/<int:problem_id>", methods=["PUT"])
def update_problem(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    data = request.get_json()
    updated = db.update_problem(problem_id, data)
    return jsonify(updated)


@app.route("/api/problems/<int:problem_id>", methods=["DELETE"])
def delete_problem(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    db.delete_problem(problem_id)
    return jsonify({"message": "Problema eliminado com sucesso"})


# --- AI: 5W1H ---

@app.route("/api/problems/<int:problem_id>/analyze", methods=["POST"])
def analyze_5w1h(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    try:
        analysis = ai.generate_5w1h(problem["title"], problem["description"], problem["area"])
        import json
        db.update_problem(problem_id, {"analysis_5w1h": json.dumps(analysis, ensure_ascii=False)})
        return jsonify(analysis)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- AI: A3 Report ---

@app.route("/api/problems/<int:problem_id>/a3", methods=["POST"])
def generate_a3(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    try:
        report = ai.generate_a3_report(problem)
        import json
        db.update_problem(problem_id, {"a3_report": json.dumps(report, ensure_ascii=False)})
        return jsonify(report)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- AI: Improvement Suggestions ---

@app.route("/api/problems/<int:problem_id>/suggestions", methods=["POST"])
def get_suggestions(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    try:
        suggestions = ai.generate_improvement_suggestions(
            problem["title"], problem["description"], problem["area"]
        )
        return jsonify(suggestions)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- AI: Suggest Actions ---

@app.route("/api/problems/<int:problem_id>/suggest_actions", methods=["POST"])
def suggest_actions(problem_id):
    problem = db.get_problem(problem_id)
    if not problem:
        return jsonify({"error": "Problema não encontrado"}), 404
    try:
        suggestions = ai.generate_action_suggestions(problem)
        return jsonify(suggestions)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# --- Actions ---

@app.route("/api/actions/upcoming", methods=["GET"])
def upcoming_actions():
    days = request.args.get("days", 7, type=int)
    return jsonify(db.get_upcoming_actions(days=days))


@app.route("/api/actions/export", methods=["GET"])
def export_actions_excel():
    """Exporta todas as ações para ficheiro .xlsx."""
    status = request.args.get("status")
    actions = db.list_actions_for_export(status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ações"

    headers = ["Problema", "Ação", "Responsável", "Deadline", "Estado", "Prioridade"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_labels = {
        "pending": "Pendente",
        "in_progress": "Em Progresso",
        "completed": "Concluído",
        "overdue": "Atrasado",
    }
    priority_labels = {
        "low": "Baixa",
        "medium": "Média",
        "high": "Alta",
        "critical": "Crítica",
    }

    for a in actions:
        ws.append([
            a.get("problem_title", "—"),
            a.get("title", "—"),
            a.get("responsible", "—"),
            a.get("deadline", "—"),
            status_labels.get(a.get("status", ""), a.get("status", "—")),
            priority_labels.get(a.get("priority", ""), a.get("priority", "—")),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"acoes_kaizen_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/api/actions", methods=["GET"])
def list_actions():
    problem_id = request.args.get("problem_id", type=int)
    status = request.args.get("status")
    actions = db.list_actions(problem_id=problem_id, status=status)
    return jsonify(actions)


@app.route("/api/actions", methods=["POST"])
def create_action():
    data = request.get_json()
    required = ["problem_id", "title", "responsible", "deadline"]
    missing = [f for f in required if not data.get(f)]
    if missing:
        return jsonify({"error": f"Campos obrigatórios em falta: {', '.join(missing)}"}), 400
    action = db.create_action(data)
    return jsonify(action), 201


@app.route("/api/actions/<int:action_id>", methods=["GET"])
def get_action(action_id):
    action = db.get_action(action_id)
    if not action:
        return jsonify({"error": "Ação não encontrada"}), 404
    return jsonify(action)


@app.route("/api/actions/<int:action_id>", methods=["PUT"])
def update_action(action_id):
    action = db.get_action(action_id)
    if not action:
        return jsonify({"error": "Ação não encontrada"}), 404
    data = request.get_json()
    updated = db.update_action(action_id, data)
    return jsonify(updated)


@app.route("/api/actions/<int:action_id>", methods=["DELETE"])
def delete_action(action_id):
    action = db.get_action(action_id)
    if not action:
        return jsonify({"error": "Ação não encontrada"}), 404
    db.delete_action(action_id)
    return jsonify({"message": "Ação eliminada com sucesso"})


# --- Seed ---

@app.route("/api/seed", methods=["POST"])
def seed():
    result = db.seed_database()
    return jsonify({
        "message": f"Base de dados populada com {result['problems']} problemas e {result['actions']} ações.",
        "problems": result["problems"],
        "actions": result["actions"],
    })


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
